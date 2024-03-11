import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from tkinter import ttk
import re
import numpy as np
from datetime import datetime, timedelta
import math

class ExcelViewer:
    def __init__(self, root):
        self.root = root
        self.root.title("Visor de Excel")
        
        self.frame = tk.Frame(root)
        self.frame.pack(fill="both", expand=True)
        
        self.table = ttk.Treeview(self.frame)
        self.table["columns"] = ()
        
        self.scroll_x = tk.Scrollbar(self.frame, orient="horizontal", command=self.table.xview)
        self.scroll_x.pack(side="bottom", fill="x")
        
        self.table.pack(fill="both", expand=True)
        self.table.configure(xscrollcommand=self.scroll_x.set)
        
        self.load_button = tk.Button(root, text="Cargar archivo Excel", command=self.load_excel)
        self.load_button.pack()
        
    def clean_data(self, value):
        if value is None or (isinstance(value, float) and np.isnan(value)):
            return ""
        return value
    
    def calculate_new_price(self, price):
        if price == "":
            return ""
        return round(price / 1.12, 2)
    
    def calculate_income(self, stock, sold):
        if stock == "" or sold == "":
            return ""
        return round(stock + sold, 2)
    
    def calculate_rotation(self, sold, income):
        if income > 0:
            return math.ceil((sold / income) * 100)
        else:
            return ""
    
    def calculate_monthly_rotation(self, sold, date):
        if sold == "" or date == "":
            return ""
        try:
            # Convertir la fecha de ingreso de formato general de Excel a una cadena
            date_str = str(date)
            # Obtener la fecha actual
            fecha_actual = datetime.now()
            # Convertir la fecha de ingreso a un objeto datetime
            fecha_ingreso = fecha_actual  # Inicializamos con la fecha actual por si hay algún problema en la conversión
            if date_str.isdigit():  # Verificamos si el valor es un número (formato general de Excel)
                days_since_1900 = int(date_str)
                # Convertir el número de días desde 1900 a una fecha
                fecha_ingreso = datetime(1899, 12, 30) + timedelta(days_since_1900 - 1)
            else:
                fecha_ingreso = datetime.strptime(date_str, "%Y-%m-%d")
            # Calcular la diferencia en meses entre la fecha actual y la fecha de ingreso
            diferencia_meses = (fecha_actual - fecha_ingreso).days / 30
            # Calcular la rotación mensual
            if diferencia_meses > 0:
                return math.ceil(sold / diferencia_meses)
            else:
                return ""
        except ValueError:
            return ""
    
    def calculate_stock_out(self, delivery_time, sold, date):
        if delivery_time == "" or sold == "" or date == "":
            return ""
        try:
            # Convertir la fecha de ingreso de formato general de Excel a una cadena
            date_str = str(date)
            # Obtener la fecha actual
            fecha_actual = datetime.now()
            # Convertir la fecha de ingreso a un objeto datetime
            fecha_ingreso = fecha_actual  # Inicializamos con la fecha actual por si hay algún problema en la conversión
            if date_str.isdigit():  # Verificamos si el valor es un número (formato general de Excel)
                days_since_1900 = int(date_str)
                # Convertir el número de días desde 1900 a una fecha
                fecha_ingreso = datetime(1899, 12, 30) + timedelta(days_since_1900 - 1)
            else:
                fecha_ingreso = datetime.strptime(date_str, "%Y-%m-%d")
            # Calcular la diferencia en días entre la fecha actual y la fecha de ingreso
            diferencia_dias = (fecha_actual - fecha_ingreso).days
            # Calcular la rotura de stock
            if diferencia_dias > 0:
                return math.ceil(delivery_time * (sold / diferencia_dias))
            else:
                return ""
        except ValueError:
            return ""
    
    def calculate_purchase_strategy(self, stock, delivery_time, sold, date):
        try:
            if not stock or not delivery_time or not sold or not date:
                return ""
            
            # Convertir la fecha de ingreso de formato general de Excel a una cadena
            date_str = str(date)
            # Obtener la fecha actual
            fecha_actual = datetime.now()
            # Convertir la fecha de ingreso a un objeto datetime
            fecha_ingreso = fecha_actual  # Inicializamos con la fecha actual por si hay algún problema en la conversión
            if date_str.isdigit():  # Verificamos si el valor es un número (formato general de Excel)
                days_since_1900 = int(date_str)
                # Convertir el número de días desde 1900 a una fecha
                fecha_ingreso = datetime(1899, 12, 30) + timedelta(days_since_1900 - 1)
            else:
                fecha_ingreso = datetime.strptime(date_str, "%Y-%m-%d")
            
            # Calcular la diferencia en días entre la fecha actual y la fecha de ingreso
            diferencia_dias = (fecha_actual - fecha_ingreso).days
            
            if diferencia_dias <= 0:
                return ""
            
            consumption_inventory = stock / (delivery_time * (sold / diferencia_dias))
            
            if consumption_inventory <= 1:
                return "Reordenar"
            elif consumption_inventory <= 1.25:
                return "Preparar"
            else:
                return ""
        except ValueError:
            return ""
    
    def calculate_holding_cost(self, cost):
        if cost == "":
            return ""
        return math.ceil(cost * (26 / 100))
    
    def calculate_reorder_quantity(self, strategy, holding_cost, stock, stock_out, sold, delivery_time, order_cost):
        if strategy == "Reordenar" or strategy == "Preparar":
            if holding_cost == "":
                return ""
            if float(holding_cost) > 0:
                cost_calculated = (stock_out - stock ) + math.sqrt((2 * sold * order_cost) / float(holding_cost))
                return math.ceil(cost_calculated)
        return ""
    
    def load_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")])
        if file_path:
            workbook = load_workbook(filename=file_path)
            sheet = workbook.active
            # Convertir nombres de columnas a cadenas válidas
            column_names = [re.sub(r'\W+', '', str(col.value)) for col in sheet[1]]
            column_names.append("NuevoPrecio")  # Agregar el nombre de la nueva columna
            column_names.append("Ingresos")  # Agregar el nombre de la nueva columna
            column_names.append("Rotacion")  # Agregar el nombre de la nueva columna
            column_names.append("RotacionMensual")  # Agregar el nombre de la nueva columna
            column_names.append("RoturaStock")  # Agregar el nombre de la nueva columna
            column_names.append("EstrategiaCompra")  # Agregar el nombre de la nueva columna
            column_names.append("CostoMantener")  # Agregar el nombre de la nueva columna
            column_names.append("CantidadReorden")  # Agregar el nombre de la nueva columna
            self.table["columns"] = column_names
            self.table.heading("#0", text="Índice")
            for i, name in enumerate(column_names):
                self.table.heading(name, text=name)
            for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column)):
                cleaned_row = [self.clean_data(cell.value) for cell in row]
                if any(cleaned_row):  # Verifica si la fila contiene algún dato no vacío
                    # Calcula el valor de la nueva columna "NuevoPrecio"
                    new_price = self.calculate_new_price(cleaned_row[3])  # Considerando que "Precio" está en la cuarta columna (índice 3)
                    cleaned_row.append(new_price)
                    # Calcula el valor de la nueva columna "Ingresos"
                    income = self.calculate_income(cleaned_row[4], cleaned_row[5])  # Considerando que "Stock" está en la quinta columna (índice 4) y "Vendido" en la sexta (índice 5)
                    cleaned_row.append(income)
                    # Calcula el valor de la nueva columna "Rotacion"
                    rotation = self.calculate_rotation(cleaned_row[5], income)  # Considerando que "Vendido" está en la sexta columna (índice 5)
                    cleaned_row.append(rotation)
                    # Calcula el valor de la nueva columna "RotacionMensual"
                    monthly_rotation = self.calculate_monthly_rotation(cleaned_row[5], cleaned_row[6])  # Considerando que "Vendido" está en la sexta columna (índice 5) y "FechaIngreso" en la séptima (índice 6)
                    cleaned_row.append(monthly_rotation)
                    # Calcula el valor de la nueva columna "RoturaStock"
                    stock_out = self.calculate_stock_out(cleaned_row[7], cleaned_row[5], cleaned_row[6])  # Considerando que "TiempoEntregaDías" está en la octava columna (índice 7)
                    cleaned_row.append(stock_out)
                    # Calcula el valor de la nueva columna "EstrategiaCompra"
                    purchase_strategy = self.calculate_purchase_strategy(cleaned_row[4], cleaned_row[7], cleaned_row[5], cleaned_row[6])  # Considerando que "Stock" está en la quinta columna (índice 4), "TiempoEntregaDías" en la octava (índice 7), "Vendido" en la sexta (índice 5) y "FechaIngreso" en la séptima (índice 6)
                    cleaned_row.append(purchase_strategy)
                    # Calcula el valor de la nueva columna "CostoMantener"
                    holding_cost = self.calculate_holding_cost(cleaned_row[2])  # Considerando que "Costo" está en la tercera columna (índice 2)
                    cleaned_row.append(holding_cost)
                    # Calcula el valor de la nueva columna "CantidadReorden"
                    reorder_quantity = self.calculate_reorder_quantity(purchase_strategy, holding_cost, cleaned_row[4], stock_out, cleaned_row[5], cleaned_row[7], cleaned_row[8])  # Considerando que "Stock" está en la quinta columna (índice 4), "RoturaStock" en la novena (índice 8), "Vendido" en la sexta (índice 5), "TiempoEntregaDías" en la octava (índice 7) y "CostoOrdenar" en la novena (índice 8)
                    cleaned_row.append(reorder_quantity)
                    # Redondear valores al inmediato superior si es necesario
                    cleaned_row = [math.ceil(value) if isinstance(value, float) and not value.is_integer() else value for value in cleaned_row]
                    self.table.insert("", "end", text=str(i+1), values=cleaned_row)
            
if __name__ == "__main__":
    root = tk.Tk()
    excel_viewer = ExcelViewer(root)
    root.mainloop()
